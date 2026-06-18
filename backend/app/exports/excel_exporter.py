from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.models import DrawingElement, CompareDiff, CompareTask
from app.services.report_service import REPORT_COLUMNS, ReportService


HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
PARAMETER_HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DEFAULT_INLINE_FONT = InlineFont(rFont="微软雅黑")
HIGHLIGHT_INLINE_FONT = InlineFont(rFont="微软雅黑", b=True, color="9C6500")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

RISK_FILLS = {
    "high": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "manual_check": PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
}

DELIVERY_COLUMN_WIDTHS = {
    "A": 16,
    "B": 8,
    "C": 18,
    "D": 14,
    "E": 34,
    "F": 34,
    "G": 42,
    "H": 28,
    "I": 32,
    "J": 10,
    "K": 12,
    "L": 34,
}

ELEMENT_REPORT_HEADERS = [
    "序号",
    "图纸类型",
    "元素名称",
    "类别",
    "原始内容",
    "单位",
    "重要性",
    "置信度",
    "需人工确认",
    "页码",
    "来源区域",
    "来源图片",
]

class ExcelExporter:
    def _style_header(self, ws, headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    def _style_cell(self, cell, wrap: bool = True) -> None:
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER

    def _auto_width(self, ws, min_width: int = 8, max_width: int = 60) -> None:
        for col in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) * 1.2, max_width))
            ws.column_dimensions[col_letter].width = max_len + 2

    def export_diffs(self, db: Session, task_id: int) -> BytesIO:
        rows = ReportService(db).build_diff_report_rows(task_id)
        task = db.query(CompareTask).filter(CompareTask.id == task_id).first()
        wb = Workbook()
        ws = wb.active
        ws.title = "差异报告"

        self._style_header(ws, REPORT_COLUMNS)
        ws.freeze_panes = "A2"

        for row_idx, report_row in enumerate(rows, 2):
            values = report_row.as_excel_values()
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                self._style_cell(cell)
            if row_idx == 2 or rows[row_idx - 2].section != rows[row_idx - 3].section:
                for col_idx in range(1, len(REPORT_COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color="D9EAF7", end_color="D9EAF7", fill_type="solid"
                    )
            if report_row.conclusion_highlights:
                conclusion_cell = ws.cell(row=row_idx, column=7)
                conclusion_cell.fill = PARAMETER_HIGHLIGHT_FILL
                conclusion_cell.value = self._rich_text_for_highlights(
                    report_row.conclusion,
                    report_row.conclusion_highlights,
                )
            ws.row_dimensions[row_idx].height = 48

        for col_letter, width in DELIVERY_COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        ws.auto_filter.ref = f"A1:L{max(ws.max_row, 1)}"

        if task:
            base_elements = self._elements_for_file(db, task.base_file_id)
            compare_elements = self._elements_for_file(db, task.compare_file_id)
            self._write_element_sheet(wb, "基准图纸元素", "基准图纸", base_elements)
            self._write_element_sheet(wb, "对比图纸元素", "对比图纸", compare_elements)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _rich_text_for_highlights(self, text: str, highlights: list[dict[str, int]]) -> CellRichText | str:
        valid_ranges = sorted(
            [
                (max(0, item["start"]), min(len(text), item["end"]))
                for item in highlights
                if item.get("end", 0) > item.get("start", 0) and item.get("start", 0) < len(text)
            ],
            key=lambda item: item[0],
        )
        if not text or not valid_ranges:
            return text

        rich_text = CellRichText()
        cursor = 0
        for start, end in valid_ranges:
            start = max(start, cursor)
            if start > cursor:
                rich_text.append(TextBlock(DEFAULT_INLINE_FONT, text[cursor:start]))
            if end > start:
                rich_text.append(TextBlock(HIGHLIGHT_INLINE_FONT, text[start:end]))
                cursor = end
        if cursor < len(text):
            rich_text.append(TextBlock(DEFAULT_INLINE_FONT, text[cursor:]))
        return rich_text if len(rich_text) > 1 else text

    def _elements_for_file(self, db: Session, file_id: int) -> list[DrawingElement]:
        return (
            db.query(DrawingElement)
            .filter(DrawingElement.file_id == file_id)
            .order_by(DrawingElement.id.asc())
            .all()
        )

    def _write_element_sheet(
        self,
        wb: Workbook,
        title: str,
        role_label: str,
        elements: list[DrawingElement],
    ) -> None:
        ws = wb.create_sheet(title=title)
        self._style_header(ws, ELEMENT_REPORT_HEADERS)
        ws.freeze_panes = "A2"

        for row_idx, elem in enumerate(elements, 2):
            page_no = elem.page.page_no if elem.page else ""
            source_area = elem.region_desc or (elem.region.region_name if elem.region else "")
            values = [
                row_idx - 1,
                role_label,
                elem.element_name,
                elem.category,
                elem.raw_value or "",
                elem.unit or "",
                elem.importance,
                f"{elem.confidence * 100:.0f}%" if elem.confidence is not None else "-",
                "是" if elem.need_manual_check else "否",
                page_no,
                source_area,
                elem.source_image_path or "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                self._style_cell(cell)

        self._auto_width(ws, min_width=8, max_width=42)
        ws.auto_filter.ref = f"A1:M{max(ws.max_row, 1)}"

    def export_elements(self, db: Session, task_id: int) -> BytesIO:
        task = db.query(CompareTask).filter(CompareTask.id == task_id).first()
        if not task:
            return BytesIO()

        elements = db.query(DrawingElement).filter(
            DrawingElement.file_id.in_([task.base_file_id, task.compare_file_id])
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "图纸元素清单"

        headers = [
            "序号", "图纸", "元素名称", "类别", "原始值",
            "单位", "重要性", "置信度", "需人工确认", "来源区域",
        ]
        self._style_header(ws, headers)

        for row_idx, elem in enumerate(elements, 2):
            file_role = "基准" if elem.file_id == task.base_file_id else "对比"
            values = [
                row_idx - 1, file_role, elem.element_name, elem.category,
                elem.raw_value or "",
                elem.unit or "", elem.importance,
                f"{elem.confidence * 100:.0f}%" if elem.confidence else "-",
                "是" if elem.need_manual_check else "否",
                elem.region_desc or "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                self._style_cell(cell)

        self._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_final_report(self, db: Session, task_id: int) -> BytesIO:
        diffs = db.query(CompareDiff).filter(CompareDiff.compare_task_id == task_id).all()
        confirmed = [d for d in diffs if d.review_status == "confirmed"]
        high_unreviewed = [d for d in diffs if d.review_status == "pending" and d.risk_level in ("high", "manual_check")]

        wb = Workbook()
        ws = wb.active
        ws.title = "审核总结"

        summary_data = [
            (1, 1, "审核总结报告", Font(name="微软雅黑", bold=True, size=14)),
            (3, 1, f"总差异数: {len(diffs)}"),
            (4, 1, f"已确认差异: {len(confirmed)}"),
            (5, 1, f"待审核高风险项: {len(high_unreviewed)}"),
            (7, 1, "高风险差异明细", HEADER_FONT),
        ]
        for row, col, text, *font in summary_data:
            cell = ws.cell(row=row, column=col, value=text)
            if font:
                cell.font = font[0]

        headers = ["序号", "差异类别", "差异说明", "基准内容", "对比内容", "审核状态"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for row_idx, diff in enumerate(diffs, 9):
            values = [
                row_idx - 8, diff.diff_category, diff.diff_summary or "",
                diff.base_content or "", diff.compare_content or "",
                diff.review_status,
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                self._style_cell(cell)

        self._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
